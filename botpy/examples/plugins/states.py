import os
import openpyxl

global stateList
def initcheck():
    try:
        global stateList
        stateList = openpyxl.load_workbook('state_list.xlsx')
        initCheckSheet = stateList.active
        initCheck = initCheckSheet['A1']
        if initCheck.value == "state_name":
            print("[Info] 状态表已经正确加载")
            return True
        else:
            print("[Info] 状态表功能异常，请查看状态表信息")
            return False
    except IOError:
        print("表格加载错误，查看表格是否放在了example目录下")

def SearchStateList(stateName):
    sheet = stateList.active
    cur_row = 1;
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=1, values_only=True):
        cur_row += 1
        for cell in row:
            print("cell:",cell)
            print("stateName:",stateName)
            if stateName == cell:
                print(f"[Info] 已查询到该状态{stateName}!\n")
                stateContent = f"\n状态：{stateName}\n"
                stateContent += f"效果：\n{sheet.cell(cur_row, 2).value}"
                return stateContent
    print(f"[Info] 查询该状态失败\"{stateName}\"!")
    stateContent = f"查询失败，该状态不存在"
    return stateContent
