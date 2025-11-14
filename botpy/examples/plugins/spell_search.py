import os
import openpyxl

global spellList
def initcheck():
    try:
        global spellList
        spellList = openpyxl.load_workbook('spell_list.xlsx')
        initCheckSheet = spellList.active
        initCheck = initCheckSheet['A1']
        if initCheck.value == "spell_name":
            print("[Info] 法表已经正确加载")
            return True
        else:
            print("[Info] 法表功能异常，请查看法表信息")
            return False
    except IOError:
        print("表格加载错误，查看表格是否放在了example目录下")

def SearchSpellList(spellName):
    sheet = spellList.active
    cur_row = 1;
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=1, values_only=True):
        cur_row += 1
        for cell in row:
            print("cell:",cell)
            print("spellName:",spellName)
            if spellName == cell:
                print(f"[Info] 已查询到该法术{spellName}!\n")
                #print(sheet.cell(cur_row, 2).value)#打印第二列
                spellContent = f"\n法术：{spellName}\n"
                spellContent += f"类型：{sheet.cell(cur_row, 2).value}\n"
                spellContent += f"消耗：{sheet.cell(cur_row, 3).value}\n"
                spellContent += f"距离：{sheet.cell(cur_row, 4).value}\n"
                spellContent += f"需求：{sheet.cell(cur_row, 5).value}\n"
                spellContent += f"持续：{sheet.cell(cur_row, 6).value}\n"
                spellContent += f"效果：\n{sheet.cell(cur_row, 7).value}\n"
                #print(spellContent)
                return spellContent
    print(f"[Info] 查询该法术失败\"{spellName}\"!")
    spellContent = f"查询失败，该法术或许不在咱们的法表里"
    return spellContent
